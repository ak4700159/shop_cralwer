from cralwer_manager import CrawlerManager
import queue
import FreeSimpleGUI as sg
import pandas as pd 
import traceback
import os 
import io
from datetime import datetime
from openpyxl.drawing.image import Image as XLImage
from utils import excel_col_width_to_pixels, pixels_to_row_height_points, autosize_text_columns
from openpyxl import Workbook
from item import ItemRow
from image import Image
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def run_all_sequential(window: sg.Window, shops: list[str], outdir: str, period: str, log_q: queue.Queue) -> None:
    """ 모든 상점에 대한 크롤링 실시 -> 순차적으로 접근, 병렬 실행 시 고쳐야될 부분이 많음 """
    try:
        ts = datetime.now().strftime("%Y-%m-%d_%H%M%S")
        combined_path = os.path.join(outdir, f"qoo10_top5_{ts}.xlsx")
        manager = CrawlerManager.get(save_path=outdir, period=period)

        # 엑셀 워크 시트 준비하기
        work_book = Workbook()
        work_sheet: Worksheet = work_book.active
        # 시트명
        work_sheet.title = "ranking"
        # 컬럼명 지정
        headers_xlsx = ["Rank", "Name", "Price(JPY)", "Price(KRW)", "Reviews",
                        "Product URL", "Shop", "Total Count", "Image"]
        work_sheet.append(headers_xlsx)
        header_fill = PatternFill("solid", fgColor="F3F6FA")
        header_font = Font(bold=True, color="1F2937")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin = Side(style="thin", color="DDDDDD")
        header_border = Border(left=thin, right=thin, top=thin, bottom=thin)

        # 헤더 행 높이/프리즈/오토필터
        work_sheet.row_dimensions[1].height = 22
        work_sheet.freeze_panes = "A2"
        work_sheet.auto_filter.ref = f"A1:I1"   # 범위는 데이터 추가 후 아래에서 다시 확장

        # 헤더 각 셀 스타일 적용 & 기본 너비(가독성 위주)
        pref_widths = [8, 40, 12, 12, 10, 50, 14, 12, 25]  # A..I
        for col_idx, _ in enumerate(headers_xlsx, start=1):
            cell = work_sheet.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = header_border
            col_letter = get_column_letter(col_idx)
            work_sheet.column_dimensions[col_letter].width = pref_widths[col_idx-1]

        for idx, shop in enumerate(shops):
            try:
                log_q.put(f"[START] {shop} 수집 시작 (period={period})")
                crawler = manager.run_shop(shop) 
                # 워크 시트에 크롤링한 데이터 전달
                _ = append_to_worksheet(work_sheet, crawler.results, crawler.images)
                window.write_event_value("-STEP_DONE-", combined_path)
                # work_book.save(combined_path)
                log_q.put(f"[DONE] {shop} 완료")
            except Exception as e:
                log_q.put("[ERROR] " + repr(e))
                log_q.put(traceback.format_exc())
                
        # 모든 상점 처리 후 통합 파일 저장
        try:
            autosize_text_columns(work_sheet, skip_letters={"I"})
            work_book.save(combined_path)
            log_q.put(f"[SAVE] 결과 저장: {combined_path}")
        except Exception as e:
            log_q.put(f"[WARN] 파일 저장 실패: {e}")
        window.write_event_value("-STEP_DONE-", combined_path)
        window.write_event_value("-ALL_DONE-", True)
    except Exception as e:
        log_q.put("[ERROR] " + repr(e))
        log_q.put(traceback.format_exc())
        window.write_event_value("-ALL_DONE-", True)

def append_to_worksheet(work_sheet: Worksheet, data_results: list[ItemRow], images: list[Image]) -> int:
    if not data_results:
        return 0

    img_col_letter = "I"
    # 이미지 열 폭은 헤더 단계에서 지정되었다고 가정(없으면 기본 지정)
    if not work_sheet.column_dimensions[img_col_letter].width:
        work_sheet.column_dimensions[img_col_letter].width = 25
    target_col_px = excel_col_width_to_pixels(work_sheet.column_dimensions[img_col_letter].width)

    # ✅ 본문 공통 스타일
    thin = Side(style="thin", color="EEEEEE")
    body_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    align_left  = Alignment(horizontal="left",  vertical="center", wrap_text=True)
    align_right = Alignment(horizontal="right", vertical="center")
    align_center= Alignment(horizontal="center",vertical="center")

    # 줄무늬(밴드) 색
    band_fill = PatternFill("solid", fgColor="FAFAFA")
    row_idx = work_sheet.max_row + 1
    for i, r in enumerate(data_results, start=1):
        # Rank
        c = work_sheet.cell(row=row_idx, column=1, value=i)
        c.alignment = align_center; c.border = body_border

        # Name
        c = work_sheet.cell(row=row_idx, column=2, value=r.name)
        c.alignment = align_left; c.border = body_border

        # Price(JPY)
        c = work_sheet.cell(row=row_idx, column=3, value=r.price_jpy)
        c.number_format = '#,##0'; c.alignment = align_right; c.border = body_border

        # Price(KRW)
        c = work_sheet.cell(row=row_idx, column=4, value=r.price_krw)
        c.number_format = '#,##0'; c.alignment = align_right; c.border = body_border

        # Reviews
        c = work_sheet.cell(row=row_idx, column=5, value=r.review_count)
        c.number_format = '#,##0'; c.alignment = align_center; c.border = body_border

        # Product URL (하이퍼링크 + 파란 밑줄)
        c = work_sheet.cell(row=row_idx, column=6, value=r.product_url)
        c.hyperlink = r.product_url
        c.style = "Hyperlink"
        c.alignment = align_left
        c.border = body_border

        # Shop
        c = work_sheet.cell(row=row_idx, column=7, value=r.shop_name)
        c.alignment = align_left; c.border = body_border

        # Total Count (문자면 그대로, 숫자면 포맷)
        c = work_sheet.cell(row=row_idx, column=8, value=r.total_count)
        try:
            float(r.total_count)
            c.number_format = '#,##0'
            c.alignment = align_right
        except Exception:
            c.alignment = align_left
        c.border = body_border

        # ✅ 밴드 채우기(가독성) — 데이터 영역 전체 셀에 적용
        if (row_idx % 2) == 0:
            for col in range(1, 9):  # A..H (이미지 I 제외)
                work_sheet.cell(row=row_idx, column=col).fill = band_fill

        # 이미지(I열)
        img_info = images[i-1]
        xlimg = XLImage(io.BytesIO(img_info.img_bytes))
        orig_w, orig_h = float(xlimg.width), float(xlimg.height)
        scale = min(1.0, target_col_px / orig_w) if orig_w > 0 else 1.0
        xlimg.width = orig_w * scale
        xlimg.height = orig_h * scale

        work_sheet.row_dimensions[row_idx].height = pixels_to_row_height_points(xlimg.height)
        work_sheet.add_image(xlimg, f"{img_col_letter}{row_idx}")

        # 이미지 셀도 테두리만
        work_sheet.cell(row=row_idx, column=9).border = body_border
        if (row_idx % 2) == 0:
            work_sheet.cell(row=row_idx, column=9).fill = band_fill

        row_idx += 1

    # ✅ 데이터 추가 후 오토필터 범위 갱신
    work_sheet.auto_filter.ref = f"A1:I{work_sheet.max_row}"
    return len(data_results)

def normalize_shop(line: str) -> str:
    line = line.strip()
    if not line:
        return ""
    if "qoo10.jp" in line:
        return line.rstrip("/").split("/")[-1]
    return line

def rows_from_one_file(path: str) -> list[list]:
    """
    방금 끝난 엑셀 파일 하나를 읽어 미리보기 테이블용 행을 반환.
    """
    try:
        if not path:
            return []
        df = pd.read_excel(path)

        # 목표 테이블 헤더 순서
        targets = ["Shop", "Name", "JPY", "KRW", "Reviews", "URL"]

        # 각 타겟 컬럼이 될 수 있는 '후보 이름'들(우선순위 순)
        candidates = {
            "Shop":    ["shop_name", "Shop"],
            "Name":    ["name", "Name"],
            "JPY":     ["price_jpy", "Price(JPY)", "JPY"],
            "KRW":     ["price_krw", "Price(KRW)", "KRW"],
            "Reviews": ["review_count", "Reviews"],
            "URL":     ["product_url", "Product URL", "URL"],
        }

        # 실제 df에 존재하는 컬럼 이름을 타겟별로 매핑
        actual = {}
        cols_set = set(df.columns)
        for tgt in targets:
            found = None
            for cand in candidates[tgt]:
                if cand in cols_set:
                    found = cand
                    break
            actual[tgt] = found

        # 존재하지 않는 나머지는 빈 문자열로 채움
        for tgt in targets:
            if actual[tgt] is None:
                df[f"__empty_{tgt}"] = ""
                actual[tgt] = f"__empty_{tgt}"

        # 타겟 순서로 DataFrame 구성 후 리스트로 변환
        out = df[[actual[t] for t in targets]].values.tolist()
        return out

    except Exception as e:
        print(f"read excel file : {e}")
        return []